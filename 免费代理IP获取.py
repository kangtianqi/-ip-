import requests  # 导入requests模块，用于发送HTTP请求
from lxml import etree  # 导入etree模块，用于解析HTML
from openpyxl import load_workbook  # 导入load_workbook模块，用于加载Excel文件
from openpyxl.workbook import Workbook  # 导入Workbook模块，用于创建新的Excel文件

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/72.0.3626.121 Safari/537.36'}

try:
    # 发起请求获取网页内容
    response = requests.get('https://www.89ip.cn/', headers=headers)
    response.encoding = 'utf-8'
    if response.status_code == 200:
        # 解析HTML
        html = etree.HTML(response.text)

        # 定位IP列表所在的表格元素
        table = html.xpath('//table[@class="layui-table"]')[0]

        # 获取所有IP行
        trs = table.xpath('//tr')[1:]
        ip_list = []

        for t in trs:
            # 获取每行的数据
            td_list = t.xpath('td/text()')
            if len(td_list) >= 2:  # 确保至少有两个元素
                # 获取IP地址和端口号
                ip = td_list[0].strip()
                port = td_list[1].strip()
                print(ip + ':' + port, '\n', end='')
                # 将IP地址和端口号保存至列表中
                ip_list.append({'ip': ip, 'port': port})

        try:
            # 判断文件是否存在
            workbook = load_workbook(filename="代理ip列表.xlsx")  # 如果文件已存在，则加载现有文件
            sheet = workbook.active
            is_new_file = False
        except FileNotFoundError:
            workbook = Workbook()  # 如果文件不存在，则创建新文件
            sheet = workbook.active
            # 设置表头
            sheet['A1'] = 'IP'
            sheet['B1'] = 'Port'
            is_new_file = True

        # 写入IP列表到Excel文件中
        start_row = sheet.max_row + 1 if not is_new_file else 2
        for i, ip_info in enumerate(ip_list):
            # 写入IP地址和端口号到对应的单元格
            sheet.cell(row=i + start_row, column=1, value=ip_info['ip'])
            sheet.cell(row=i + start_row, column=2, value=ip_info['port'])

        # 保存Excel文件
        workbook.save(filename="代理ip列表.xlsx")

except requests.exceptions.RequestException as e:
    print('请求异常:', e)
